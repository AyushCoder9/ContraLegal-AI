from collections import Counter
import json
import os
import pickle
from typing import List, Literal

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer

from .inference_engine import RiskScorer

MODEL_PATH = "models/model.pkl"
VECTORIZER_PATH = "models/vectorizer.pkl"


def load_model():
    if not os.path.exists(MODEL_PATH) or not os.path.exists(VECTORIZER_PATH):
        return None, None
    with open(VECTORIZER_PATH, "rb") as f:
        vectorizer = pickle.load(f)
    with open(MODEL_PATH, "rb") as f:
        model = pickle.load(f)
    return vectorizer, model


def _get_ml_scores(sentences: List[str], vectorizer, model) -> List[float]:
    tfidf_matrix = vectorizer.transform(sentences)
    if hasattr(model, "predict_proba"):
        return model.predict_proba(tfidf_matrix)[:, 1].tolist()
    return [float(p) for p in model.predict(tfidf_matrix)]


def predict_hybrid(sentences: List[str], vectorizer, model) -> pd.DataFrame:
    scorer = RiskScorer()
    ml_scores = _get_ml_scores(sentences, vectorizer, model)
    results = [
        scorer.score_clause(clause, ml_score)
        for clause, ml_score in zip(sentences, ml_scores)
    ]
    df = pd.DataFrame(results)
    df = df.sort_values("final_score", ascending=False).reset_index(names="clause_index")
    return df


def compute_contract_risk(df: pd.DataFrame) -> dict:
    overall = round(float(df["final_score"].mean()), 4)

    if overall >= 0.5:
        label = "High Risk"
    elif overall >= 0.3:
        label = "Medium Risk"
    else:
        label = "Low Risk"

    highest_clause_row = df.loc[df["final_score"].idxmax()]
    high_count = int((df["final_score"] >= 0.5).sum())
    medium_count = int(((df["final_score"] >= 0.3) & (df["final_score"] < 0.5)).sum())
    low_count = int((df["final_score"] < 0.3).sum())

    all_kws: List[str] = []
    for flags in df["keyword_flags"]:
        if isinstance(flags, list):
            all_kws.extend(flags)
    kw_counts = Counter(all_kws)
    top_keyword_counts = [{"term": kw, "count": cnt} for kw, cnt in kw_counts.most_common(10)]

    return {
        "overall_score": overall,
        "risk_label": label,
        "highest_risk_clause": highest_clause_row["clause_text"],
        "highest_risk_score": round(float(highest_clause_row["final_score"]), 4),
        "high_risk_count": high_count,
        "medium_risk_count": medium_count,
        "low_risk_count": low_count,
        "top_keywords": [x["term"] for x in top_keyword_counts[:5]],
        "keyword_frequency": top_keyword_counts,
    }


def export_results(
    df: pd.DataFrame,
    output_path: str,
    fmt: Literal["json", "csv"] = "json",
) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    export_df = df.copy()
    for col in ("keyword_flags", "keyword_matches"):
        if col in export_df.columns:
            export_df[col] = export_df[col].apply(
                lambda v: ", ".join(str(x) for x in v) if isinstance(v, list) else v
            )
    if fmt == "csv":
        export_df.to_csv(output_path, index=False)
    else:
        records = export_df.to_dict(orient="records")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
    return os.path.abspath(output_path)


def build_xlsx_export(df: pd.DataFrame, contract_risk: dict) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    output = io.BytesIO()

    summary_rows = [
        ("Overall Risk Label", contract_risk["risk_label"]),
        ("Overall Risk Score", f"{contract_risk['overall_score']:.0%}"),
        ("Total Clauses", contract_risk["high_risk_count"] + contract_risk["medium_risk_count"] + contract_risk["low_risk_count"]),
        ("High Risk Clauses", contract_risk["high_risk_count"]),
        ("Medium Risk Clauses", contract_risk["medium_risk_count"]),
        ("Low Risk Clauses", contract_risk["low_risk_count"]),
        ("Highest Risk Score", f"{contract_risk['highest_risk_score']:.0%}"),
        ("Highest Risk Clause", contract_risk["highest_risk_clause"][:300]),
        ("", ""),
        ("Top Risk Keywords (by frequency)", ""),
    ]
    for entry in contract_risk.get("keyword_frequency", []):
        summary_rows.append((f"  {entry['term']}", entry["count"]))

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 60

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    label_font = Font(bold=True)

    ws_summary.append(["ContraLegal-AI · Contract Risk Summary", ""])
    ws_summary["A1"].font = Font(bold=True, size=14, color="6366F1")
    ws_summary["A1"].alignment = Alignment(horizontal="left")
    ws_summary.append(["", ""])

    ws_summary.append(["Field", "Value"])
    ws_summary["A3"].fill = header_fill
    ws_summary["A3"].font = header_font
    ws_summary["B3"].fill = header_fill
    ws_summary["B3"].font = header_font

    risk_colors = {"High Risk": "FECACA", "Medium Risk": "FFEDD5", "Low Risk": "BBF7D0"}

    for i, (field, value) in enumerate(summary_rows, start=4):
        ws_summary.append([field, value])
        if field == "Overall Risk Label":
            color = risk_colors.get(str(value), "FFFFFF")
            ws_summary[f"B{i}"].fill = PatternFill("solid", fgColor=color)
            ws_summary[f"B{i}"].font = Font(bold=True)
        elif field and not field.startswith(" "):
            ws_summary[f"A{i}"].font = label_font

    ws_detail = wb.create_sheet("Clause Detail")
    detail_cols = ["clause_index", "clause_text", "risk_label", "final_score", "ml_score", "keyword_flags"]
    export_df = df.drop(columns=["keyword_matches"], errors="ignore").copy()
    if "keyword_flags" in export_df.columns:
        export_df["keyword_flags"] = export_df["keyword_flags"].apply(
            lambda v: ", ".join(v) if isinstance(v, list) else v
        )

    present_cols = [c for c in detail_cols if c in export_df.columns]
    ws_detail.append(present_cols)
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font

    risk_row_colors = {"High Risk": "FECACA", "Medium Risk": "FFEDD5", "Low Risk": "BBF7D0"}
    for _, row in export_df[present_cols].iterrows():
        ws_detail.append([row[c] for c in present_cols])
        row_idx = ws_detail.max_row
        label = row.get("risk_label", "")
        fill_color = risk_row_colors.get(label, "FFFFFF")
        for col_idx in range(1, len(present_cols) + 1):
            ws_detail.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=fill_color)

    ws_detail.column_dimensions["B"].width = 80
    ws_detail.column_dimensions["C"].width = 14
    ws_detail.column_dimensions["F"].width = 40

    wb.save(output)
    return output.getvalue()


def summarize_contract(clauses: List[str], top_n: int = 5) -> List[dict]:
    if len(clauses) <= top_n:
        return [{"clause_text": c, "importance_score": 1.0} for c in clauses]
    vec = TfidfVectorizer(stop_words="english")
    tfidf_matrix = vec.fit_transform(clauses)
    scores = np.asarray(tfidf_matrix.sum(axis=1)).flatten()
    top_indices = scores.argsort()[::-1][:top_n]
    return [
        {"clause_text": clauses[i], "importance_score": round(float(scores[i]), 4)}
        for i in sorted(top_indices)
    ]


def generate_clause_clusters(
    sentences: List[str], vectorizer, num_clusters: int = 4
) -> tuple[list[int], dict]:
    if len(sentences) < num_clusters:
        num_clusters = max(1, len(sentences))
    tfidf_matrix = vectorizer.transform(sentences)
    kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)
    clusters = kmeans.fit_predict(tfidf_matrix)
    order_centroids = kmeans.cluster_centers_.argsort()[:, ::-1]
    terms = vectorizer.get_feature_names_out()
    cluster_headings = {
        i: " / ".join(terms[ind].title() for ind in order_centroids[i, :3])
        for i in range(num_clusters)
    }
    return clusters.tolist(), cluster_headings
